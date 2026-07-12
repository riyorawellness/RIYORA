"""Batch 4 — User 360 export + new report types (payouts/pending/revenue)."""
import io
import os
from pathlib import Path

import pytest
import requests
from openpyxl import load_workbook

_env = Path("/app/frontend/.env")
for _ln in _env.read_text().splitlines():
    if _ln.startswith("REACT_APP_BACKEND_URL"):
        os.environ["REACT_APP_BACKEND_URL"] = _ln.split("=", 1)[1].strip().strip('"')

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"
ADMIN_MOBILE = "9999999999"
ADMIN_PASSWORD = "Admin@12345"


@pytest.fixture(scope="module")
def admin_h():
    r = requests.post(
        f"{API}/admin/login", json={"mobile": ADMIN_MOBILE, "password": ADMIN_PASSWORD}
    )
    return {
        "Authorization": f"Bearer {r.json()['tokens']['access_token']}",
        "Content-Type": "application/json",
    }


@pytest.fixture(scope="module")
def sample_user(admin_h):
    """Pick any real user (not the company root) to test 360 on."""
    r = requests.get(f"{API}/admin/users?page=1&page_size=10", headers=admin_h)
    items = [u for u in r.json().get("items", []) if u["membership_id"] != "RW000000"]
    assert items, "No users to test with"
    return items[0]


class TestNewReportTypes:
    @pytest.mark.parametrize("report_type", ["payouts", "pending_payments", "revenue_summary"])
    def test_list_endpoint_shape(self, admin_h, report_type):
        r = requests.get(
            f"{API}/admin/reports/{report_type}?page_size=5", headers=admin_h
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["report_type"] == report_type
        assert isinstance(d["columns"], list)
        assert isinstance(d["items"], list)
        # Column shape sanity
        assert all("key" in c and "label" in c for c in d["columns"])

    def test_revenue_summary_buckets_monthly(self, admin_h):
        r = requests.get(f"{API}/admin/reports/revenue_summary", headers=admin_h)
        rows = r.json()["items"]
        # Each row must have all keys we advertise
        for row in rows:
            assert "period" in row and len(row["period"]) == 7  # YYYY-MM
            for k in ("purchases", "razorpay_amount", "qr_amount", "gst_amount", "total"):
                assert k in row

    def test_revenue_summary_yearly_via_level_toggle(self, admin_h):
        # `level=1` toggles yearly buckets in the builder.
        r = requests.get(
            f"{API}/admin/reports/revenue_summary?level=1", headers=admin_h
        )
        rows = r.json()["items"]
        for row in rows:
            assert len(row["period"]) == 4  # YYYY

    @pytest.mark.parametrize("fmt", ["csv", "excel", "pdf"])
    def test_new_reports_export(self, admin_h, fmt):
        r = requests.get(
            f"{API}/admin/reports/revenue_summary/export?fmt={fmt}",
            headers=admin_h,
        )
        assert r.status_code == 200
        assert len(r.content) > 100


class TestUser360:
    def test_user_360_json(self, admin_h, sample_user):
        r = requests.get(
            f"{API}/admin/reports/user-360/{sample_user['membership_id']}",
            headers=admin_h,
        )
        assert r.status_code == 200
        d = r.json()
        for section in (
            "profile", "sponsor", "meter", "aggregates",
            "downline", "payments", "programs", "commissions",
            "activity", "logins", "payouts",
        ):
            assert section in d
        # Aggregates present
        assert set(d["aggregates"].keys()) >= {
            "total_paid", "total_commission_earned",
            "downline_count", "purchases_count", "programs_touched",
        }

    def test_user_360_unknown_user_404(self, admin_h):
        r = requests.get(
            f"{API}/admin/reports/user-360/RW999999", headers=admin_h
        )
        assert r.status_code == 404

    def test_user_360_excel_export(self, admin_h, sample_user):
        r = requests.get(
            f"{API}/admin/reports/user-360/{sample_user['membership_id']}/export",
            headers=admin_h,
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml"
        )
        wb = load_workbook(io.BytesIO(r.content))
        # Every 360 export must have these 8 sheets.
        expected = {"Profile", "Downline", "Payments", "Programs",
                    "Commissions", "Payouts", "Activity", "Logins"}
        assert expected.issubset(set(wb.sheetnames))

        # Profile sheet: field/value pairs — should have Membership ID row.
        ws = wb["Profile"]
        rows = [(ws.cell(r, 1).value, ws.cell(r, 2).value) for r in range(2, ws.max_row + 1)]
        keys = [k for k, _ in rows]
        assert "Membership ID" in keys
        assert "Total paid" in keys

    def test_user_360_requires_admin(self, sample_user):
        r = requests.get(
            f"{API}/admin/reports/user-360/{sample_user['membership_id']}"
        )
        assert r.status_code in (401, 403)
