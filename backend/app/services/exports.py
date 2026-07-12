"""Generic tabular exporters — CSV, Excel (xlsx), PDF.

Used by Phase 8 admin/user reports. The public API is `export_table()` which
accepts a list of dict rows + column definitions and returns
`(bytes, media_type, filename)`.

Columns spec:
    [{"key": "created_at", "label": "Date", "width": 24},
     {"key": "total",      "label": "Amount", "type": "money"},
     {"key": "status",     "label": "Status"}, ...]

Supported column types: str (default), money, date, int, bool.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any, Iterable, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ExportFormat = Literal["csv", "excel", "pdf"]

ROYAL = colors.HexColor("#0B1A5B")
GOLD = colors.HexColor("#B08A3E")
INK = colors.HexColor("#1F2937")
MUTED = colors.HexColor("#6B7280")


def _fmt_value(value: Any, col_type: str | None) -> str:
    if value is None or value == "":
        return "—"
    if col_type == "money":
        try:
            return f"Rs. {float(value):,.2f}"
        except Exception:
            return str(value)
    if col_type == "date":
        try:
            s = str(value)
            return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%d %b %Y")
        except Exception:
            return str(value)[:10]
    if col_type == "datetime":
        try:
            s = str(value)
            return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%d %b %Y %H:%M")
        except Exception:
            return str(value)[:16]
    if col_type == "int":
        try:
            return f"{int(value):,}"
        except Exception:
            return str(value)
    if col_type == "bool":
        return "Yes" if value else "No"
    return str(value)


def _raw_value(value: Any, col_type: str | None) -> Any:
    if value is None:
        return ""
    if col_type in ("money", "int"):
        try:
            return float(value) if col_type == "money" else int(value)
        except Exception:
            return value
    return value


# ---------- CSV ---------------------------------------------------------------


def to_csv(columns: list[dict], rows: Iterable[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([c["label"] for c in columns])
    for r in rows:
        writer.writerow([_raw_value(r.get(c["key"]), c.get("type")) for c in columns])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


# ---------- Excel -------------------------------------------------------------


def to_excel(columns: list[dict], rows: Iterable[dict], sheet_name: str = "Report") -> bytes:
    return to_excel_multi_sheet([(sheet_name, columns, list(rows))])


def to_excel_multi_sheet(
    sheets: list[tuple[str, list[dict], list[dict]]],
) -> bytes:
    """Write multiple sheets in ONE workbook.
    Each entry is (sheet_name, columns, rows).
    """
    wb = Workbook()
    # Remove the default active sheet — we'll add ours.
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B1A5B")
    center = Alignment(horizontal="center", vertical="center")

    for sheet_name, columns, rows in sheets:
        safe_name = (sheet_name or "Sheet")[:31] or "Sheet"
        ws = wb.create_sheet(safe_name)
        for idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx, value=col["label"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[get_column_letter(idx)].width = col.get("width", 18)
        for r_i, row in enumerate(rows, start=2):
            for c_i, col in enumerate(columns, start=1):
                v = _raw_value(row.get(col["key"]), col.get("type"))
                cell = ws.cell(row=r_i, column=c_i, value=v)
                if col.get("type") == "money":
                    cell.number_format = '"₹"#,##0.00'
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- PDF ---------------------------------------------------------------


def _pdf_styles():
    s = getSampleStyleSheet()
    s.add(ParagraphStyle("H1", parent=s["Heading1"], fontSize=18, leading=22, textColor=ROYAL))
    s.add(
        ParagraphStyle(
            "Eyebrow", parent=s["Normal"], fontSize=8, leading=10, textColor=GOLD,
            fontName="Helvetica-Bold",
        )
    )
    s.add(ParagraphStyle("Muted", parent=s["Normal"], fontSize=8, textColor=MUTED))
    return s


def to_pdf(
    columns: list[dict],
    rows: list[dict],
    title: str,
    subtitle: str | None = None,
    summary_lines: list[str] | None = None,
) -> bytes:
    styles = _pdf_styles()
    buf = io.BytesIO()
    pagesize = landscape(A4) if len(columns) > 6 else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )
    flow: list[Any] = [
        Paragraph("R I Y O R A &nbsp; W E L L N E S S", styles["Eyebrow"]),
        Paragraph(title, styles["H1"]),
    ]
    if subtitle:
        flow.append(Paragraph(subtitle, styles["Muted"]))
    flow.append(
        Paragraph(
            f"Generated {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')} · "
            f"{len(rows)} row(s)",
            styles["Muted"],
        )
    )
    if summary_lines:
        flow.append(Spacer(1, 6))
        for line in summary_lines:
            flow.append(Paragraph(line, styles["Muted"]))
    flow.append(Spacer(1, 10))

    # Data table
    header = [c["label"] for c in columns]
    data = [header]
    for r in rows:
        data.append([_fmt_value(r.get(c["key"]), c.get("type")) for c in columns])
    if len(data) == 1:
        data.append(["—"] * len(header))

    # column widths
    page_w = pagesize[0] - (24 * mm)
    total = sum(c.get("width", 20) for c in columns)
    col_widths = [(c.get("width", 20) / total) * page_w for c in columns]

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), ROYAL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]
        )
    )
    flow.append(tbl)
    doc.build(flow)
    return buf.getvalue()


# ---------- Facade ------------------------------------------------------------


MEDIA = {
    "csv": "text/csv",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}
EXT = {"csv": "csv", "excel": "xlsx", "pdf": "pdf"}


def export_table(
    fmt: ExportFormat,
    columns: list[dict],
    rows: list[dict],
    title: str,
    filename_stem: str,
    subtitle: str | None = None,
    summary_lines: list[str] | None = None,
) -> tuple[bytes, str, str]:
    """Return (bytes, media_type, filename)."""
    if fmt == "csv":
        content = to_csv(columns, rows)
    elif fmt == "excel":
        content = to_excel(columns, rows, sheet_name=title[:30])
    elif fmt == "pdf":
        content = to_pdf(columns, rows, title=title, subtitle=subtitle, summary_lines=summary_lines)
    else:
        raise ValueError(f"Unknown format: {fmt}")
    return content, MEDIA[fmt], f"{filename_stem}.{EXT[fmt]}"
